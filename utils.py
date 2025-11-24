import os
from gtts import gTTS
from openpyxl import Workbook, load_workbook

EXCEL_FILE = "calls.xlsx"

def ensure_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "call_sid", "phone_number", "name", "loan_amount",
            "tenure_months", "interest_rate", "employment_status", "confirmation"
        ])
        wb.save(EXCEL_FILE)

def log_call(call_sid, phone, name, loan, tenure, rate, emp_status, confirm):
    ensure_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([call_sid, phone, name, loan, tenure, rate, emp_status, confirm])
    wb.save(EXCEL_FILE)

def get_tts_cached(text, filename):
    os.makedirs("responses", exist_ok=True)
    path = os.path.join("responses", filename)
    if not os.path.exists(path):
        tts = gTTS(text=text, lang="hi")
        tts.save(path)
    return path
